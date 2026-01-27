"""
Excel-Import für Schichtpläne
"""
from openpyxl import load_workbook
from arbeitszeit.models import Mitarbeiter
from schichtplan.models import Schichtplan, Schicht


class SchichtplanImporter:
    
    def import_excel_mit_zuordnung(self, file_path, schichtplan):
        """
        Excel-Import mit MA-Zuordnung
        """
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Spalte 3+ = MA1, MA2, ...
        for col_idx in range(3, ws.max_column + 1):
            ma_nummer = col_idx - 2
            
            try:
                mitarbeiter = Mitarbeiter.objects.get(
                    schichtplan_kennung=f'MA{ma_nummer}'
                )
                # ... Import-Logik ...
                
            except Mitarbeiter.DoesNotExist:
                # Warnung
                pass